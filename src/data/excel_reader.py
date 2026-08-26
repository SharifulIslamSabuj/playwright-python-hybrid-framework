"""Excel-based DDT dataset reader.

Reads `tests/data/registration_test_data.xlsx` (worksheet
`RegistrationData`) into the same shape `src/data/users.py`'s
`REGISTRATION_PROFILES` previously held as a Python literal — the Excel
file is now the authoritative DDT source, this module is the only place
that knows how to read it. Contains no test-data business logic (field
defaults, generation, mode selection) — that stays in `src/data/users.py`,
per docs/11-Framework-Architecture.md §22's separation rule (data-source
reading vs. data-shape business logic).

Uses openpyxl (already the project's only spreadsheet dependency) in
read-only mode — no write-back, no formula evaluation needed.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

WORKSHEET_NAME = "RegistrationData"

# The columns every row must supply. `test_data_id`/`profile_name` are
# metadata (traceability/pytest-ID sourcing) and are returned alongside the
# profile fields, not folded into the NewUserPayload-shaped dict directly —
# callers pick what they need (src/data/users.py strips metadata before
# building a NewUserPayload).
REQUIRED_COLUMNS: tuple[str, ...] = (
    "test_data_id",
    "profile_name",
    "title",
    "birth_date",
    "birth_month",
    "birth_year",
    "firstname",
    "lastname",
    "company",
    "address1",
    "address2",
    "country",
    "zipcode",
    "state",
    "city",
    "mobile_number",
)


class ExcelDataError(Exception):
    """Raised for any malformed/incomplete Excel DDT dataset — fails fast
    with a specific, actionable message rather than letting a bad row reach
    test execution as a confusing downstream KeyError/AssertionError."""


def read_registration_profiles(path: Path | str = "tests/data/registration_test_data.xlsx") -> list[dict[str, str]]:
    """Reads every data row from the `RegistrationData` worksheet into a
    list of dicts (row order preserved, so `profile_index` stays a stable,
    predictable reference into the returned list — the same contract
    `REGISTRATION_PROFILES[profile_index]` had as a Python list).

    Raises `ExcelDataError` for: a missing file, a missing worksheet,
    missing/misordered required columns, or any row with an empty required
    value. Always closes the workbook, including on error.
    """
    resolved_path = Path(path)
    if not resolved_path.is_file():
        raise ExcelDataError(f"Excel DDT source not found: {resolved_path}")

    try:
        workbook = load_workbook(filename=resolved_path, read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises several distinct exception types for a corrupt/invalid file
        raise ExcelDataError(f"Could not open Excel DDT source {resolved_path}: {exc}") from exc

    try:
        if WORKSHEET_NAME not in workbook.sheetnames:
            raise ExcelDataError(
                f"Worksheet '{WORKSHEET_NAME}' not found in {resolved_path} "
                f"(found: {workbook.sheetnames})"
            )
        worksheet = workbook[WORKSHEET_NAME]

        rows_iter = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise ExcelDataError(f"{resolved_path}::{WORKSHEET_NAME} has no header row") from None

        headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
        missing_columns = [col for col in REQUIRED_COLUMNS if col not in headers]
        if missing_columns:
            raise ExcelDataError(
                f"{resolved_path}::{WORKSHEET_NAME} is missing required column(s): "
                f"{missing_columns} (found columns: {headers})"
            )

        profiles: list[dict[str, str]] = []
        for row_number, row in enumerate(rows_iter, start=2):  # row 1 is the header
            if row is None or all(cell is None for cell in row):
                continue  # skip a genuinely blank trailing row, not a data row
            row_dict = dict(zip(headers, row, strict=False))
            record: dict[str, str] = {}
            for column in REQUIRED_COLUMNS:
                value = row_dict.get(column)
                if value is None or str(value).strip() == "":
                    raise ExcelDataError(
                        f"{resolved_path}::{WORKSHEET_NAME} row {row_number}: "
                        f"required column '{column}' is empty"
                    )
                record[column] = str(value).strip()
            profiles.append(record)

        if not profiles:
            raise ExcelDataError(f"{resolved_path}::{WORKSHEET_NAME} has no data rows")

        return profiles
    finally:
        workbook.close()
